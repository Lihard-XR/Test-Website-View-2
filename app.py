import os
import time
import threading
import datetime as dt
from io import BytesIO

import pyodbc
from dotenv import load_dotenv
from flask import (
    Flask,
    jsonify,
    render_template,
    send_from_directory,
    request,
    redirect,
    send_file,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# =========================================================
# 1. App Initialization
# =========================================================
load_dotenv()
app = Flask(__name__)


# =========================================================
# 2. Runtime Config / Cache
# =========================================================
# 대시보드 최신 상태를 메모리에 유지하는 캐시
# /api/machines 에서는 이 캐시를 그대로 반환해 DB 부하를 줄인다.
_CACHE = {
    "ts": 0.0,
    "data": None,
    "error": None,
    "running": False,
}

# machine_status 스키마 캐시
_MACHINE_STATUS_SCHEMA_CACHE = {
    "cols": None,
    "ts": 0.0,
}

# 공구관리 조회 캐시
_TOOL_STATUS_CACHE = {
    # key: (start_date, end_date, machine, keyword)
    # value: {"ts": float, "rows": list}
}

# DB 스냅샷 갱신 주기
REFRESH_SEC = 2.0

# 에뮬레이터 replay 기본 step
REPLAY_STEP_SEC = 5


# =========================================================
# 3. Environment / Utility Helpers
# =========================================================
def env(name: str, default=None, required=True):
    """
    환경변수 조회 헬퍼.

    required=True 인데 값이 비어 있으면 즉시 예외를 발생시켜
    서버 시작 초기에 설정 누락을 확인할 수 있게 한다.
    """
    v = os.getenv(name, default)
    if required and (v is None or str(v).strip() == ""):
        raise RuntimeError(f"Missing env: {name}")
    return v


def parse_local_datetime(value: str):
    """
    문자열을 datetime 객체로 변환한다.

    프론트 입력 포맷이 약간 달라도 받을 수 있도록
    여러 형식을 순차적으로 허용한다.
    """
    if not value:
        return None

    value = value.strip()
    fmts = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
    ]

    for fmt in fmts:
        try:
            return dt.datetime.strptime(value, fmt)
        except ValueError:
            pass

    try:
        return dt.datetime.fromisoformat(value)
    except Exception:
        return None


def safe_int(v, default=0):
    """안전한 int 변환."""
    try:
        if v is None or v == "":
            return default
        return int(float(v))
    except Exception:
        return default


def safe_float(v, default=0.0):
    """안전한 float 변환."""
    try:
        if v is None or v == "":
            return default
        return float(v)
    except Exception:
        return default


def format_ts(value):
    """
    datetime 타입이면 문자열로 통일한다.
    API 응답에서 직렬화 일관성을 맞추기 위한 함수.
    """
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


def row_from_cursor_fetchone(cur, row):
    """
    pyodbc row 1건을 dict로 변환한다.
    """
    if not row:
        return None
    cols = [c[0] for c in cur.description]
    return dict(zip(cols, row))


def rows_from_cursor_fetchall(cur, rows):
    """
    pyodbc row 여러 건을 dict list로 변환한다.
    """
    cols = [c[0] for c in cur.description]
    return [dict(zip(cols, r)) for r in rows]

def sql_ident(name: str):
    """
    SQL 식별자 컬럼명을 안전하게 감싼다.
    """
    if not name:
        return None
    return f"[{str(name).replace(']', ']]')}]"


def get_machine_status_columns(conn, refresh: bool = False):
    """
    dbo.machine_status 테이블의 실제 컬럼 목록을 조회한다.
    컬럼명이 현장 DB마다 다를 수 있어 동적으로 탐지하기 위해 사용한다.
    """
    now = time.time()

    if (
        not refresh
        and _MACHINE_STATUS_SCHEMA_CACHE["cols"] is not None
        and (now - _MACHINE_STATUS_SCHEMA_CACHE["ts"]) < 300
    ):
        return _MACHINE_STATUS_SCHEMA_CACHE["cols"]

    cur = conn.cursor()
    cur.execute("""
        SELECT c.name
        FROM sys.columns c
        JOIN sys.objects o
          ON c.object_id = o.object_id
        JOIN sys.schemas s
          ON o.schema_id = s.schema_id
        WHERE s.name = 'dbo'
          AND o.name = 'machine_status'
        ORDER BY c.column_id
    """)

    cols = [str(r[0]) for r in cur.fetchall()]
    _MACHINE_STATUS_SCHEMA_CACHE["cols"] = cols
    _MACHINE_STATUS_SCHEMA_CACHE["ts"] = now
    return cols


def pick_existing_column(cols, candidates):
    """
    후보 컬럼명 중 실제 존재하는 첫 번째 컬럼을 반환한다.
    """
    lowered = {c.lower(): c for c in cols}
    for cand in candidates:
        found = lowered.get(str(cand).lower())
        if found:
            return found
    return None


def parse_date_only(value: str):
    """
    YYYY-MM-DD 문자열을 date로 변환.
    """
    if not value:
        return None
    return dt.datetime.strptime(value.strip(), "%Y-%m-%d").date()


def make_tool_warning_text(life_pct: float, raw_warning: str = ""):
    """
    경고 문구를 통일한다.
    DB에 별도 경고 컬럼이 있으면 우선 사용하고,
    없으면 수명 기준으로 간단히 산출한다.
    """
    raw_warning = str(raw_warning or "").strip()
    if raw_warning:
        return raw_warning

    pct = safe_float(life_pct, 0)

    if pct >= 95:
        return "교체필요"
    if pct >= 80:
        return "주의"
    return ""


def make_tool_cache_key(start_date: str, end_date: str, machine: str = "", keyword: str = ""):
    return (
        (start_date or "").strip(),
        (end_date or "").strip(),
        (machine or "").strip(),
        (keyword or "").strip().lower(),
    )


def query_tool_status_machine_list():
    """
    공구관리 현황 설비 드롭다운용 목록.
    현재 캐시된 설비 목록을 재사용한다.
    """
    data = _CACHE.get("data") or {}
    rows = data.get("rows") or []

    seen = set()
    result = []

    for row in rows:
        ip = str(row.get("ip") or "").strip()
        name = str(row.get("name") or ip).strip()

        if not ip or ip in seen:
            continue

        seen.add(ip)
        result.append({
            "ip": ip,
            "name": name or ip,
        })

    result.sort(key=lambda x: (x["name"], x["ip"]))
    return result


# =========================================================
# 4. Database Connection
# =========================================================
def get_conn():
    """
    MSSQL 연결 생성.
    """
    return pyodbc.connect(
        f"DRIVER={{{env('MSSQL_DRIVER')}}};"
        f"SERVER={env('MSSQL_HOST')},{env('MSSQL_PORT')};"
        f"DATABASE={env('MSSQL_DB')};"
        f"UID={env('MSSQL_USER')};"
        f"PWD={env('MSSQL_PASS')};"
        f"Encrypt=yes;"
        f"TrustServerCertificate=yes;"
        f"Connection Timeout=10;"
    )


# =========================================================
# 5. Cache Access Helpers
# =========================================================
def get_cached_machine_row(ip: str):
    """
    캐시된 최신 스냅샷에서 특정 설비 row를 반환한다.
    """
    data = _CACHE.get("data") or {}
    rows = data.get("rows") or []

    for row in rows:
        if str(row.get("ip") or "") == str(ip):
            return dict(row)
    return None


# =========================================================
# 6. Query Functions - Dashboard / Machine Snapshot
# =========================================================
def query_today_output_by_ip(conn, target_time: dt.datetime | None = None) -> dict:
    """
    기준 시각까지의 당일 누적 생산량(part_count)을 ip별로 계산한다.

    계산 기준:
    1) 전일 마지막 part_count를 baseline으로 우선 사용
    2) 전일 값이 없으면 당일 첫 비-OFFLINE row의 part_count를 baseline으로 사용
    3) 그것도 없으면 0 사용

    생산량:
    - current_count - baseline_count
    - 음수면 카운터 리셋으로 보고 current_count 사용
    """
    out = {}
    cur = conn.cursor()

    base_time = target_time or dt.datetime.now()
    today_start = dt.datetime.combine(base_time.date(), dt.time(0, 0, 0))
    prev_day_start = today_start - dt.timedelta(days=1)

    cur.execute("""
        WITH prev_last AS (
            SELECT
                ip,
                name,
                ISNULL(part_count, 0) AS prev_count,
                ROW_NUMBER() OVER (
                    PARTITION BY ip
                    ORDER BY [timestamp] DESC, id DESC
                ) AS rn
            FROM dbo.machine_status
            WHERE [timestamp] >= ?
              AND [timestamp] < ?
              AND ip IS NOT NULL
        ),
        today_first_online AS (
            SELECT
                ip,
                name,
                ISNULL(part_count, 0) AS first_online_count,
                ROW_NUMBER() OVER (
                    PARTITION BY ip
                    ORDER BY [timestamp] ASC, id ASC
                ) AS rn
            FROM dbo.machine_status
            WHERE [timestamp] >= ?
              AND [timestamp] <= ?
              AND ip IS NOT NULL
              AND ISNULL(status, '') NOT IN ('OFFLINE')
        ),
        today_last AS (
            SELECT
                ip,
                name,
                ISNULL(part_count, 0) AS current_count,
                ROW_NUMBER() OVER (
                    PARTITION BY ip
                    ORDER BY [timestamp] DESC, id DESC
                ) AS rn
            FROM dbo.machine_status
            WHERE [timestamp] >= ?
              AND [timestamp] <= ?
              AND ip IS NOT NULL
        ),
        merged AS (
            SELECT
                COALESCE(tl.ip, pl.ip, fo.ip) AS ip,
                COALESCE(tl.name, pl.name, fo.name, COALESCE(tl.ip, pl.ip, fo.ip)) AS name,
                pl.prev_count,
                fo.first_online_count,
                tl.current_count
            FROM (SELECT ip, name, prev_count FROM prev_last WHERE rn = 1) pl
            FULL OUTER JOIN (SELECT ip, name, first_online_count FROM today_first_online WHERE rn = 1) fo
              ON pl.ip = fo.ip
            FULL OUTER JOIN (SELECT ip, name, current_count FROM today_last WHERE rn = 1) tl
              ON COALESCE(pl.ip, fo.ip) = tl.ip
        )
        SELECT
            ip,
            name,
            ISNULL(prev_count, 0) AS prev_count,
            ISNULL(first_online_count, 0) AS first_online_count,
            ISNULL(current_count, 0) AS current_count
        FROM merged
        WHERE ip IS NOT NULL
    """, (
        prev_day_start, today_start,
        today_start, base_time,
        today_start, base_time,
    ))

    for ip, name, prev_count, first_online_count, current_count in cur.fetchall():
        ip = str(ip)
        prev_count = safe_int(prev_count, 0)
        first_online_count = safe_int(first_online_count, 0)
        current_count = safe_int(current_count, 0)

        # baseline 우선순위:
        # 1. 전일 마지막 값
        # 2. 당일 첫 비-OFFLINE 값
        # 3. 0
        if prev_count > 0:
            baseline = prev_count
        elif first_online_count > 0:
            baseline = first_online_count
        else:
            baseline = 0

        if current_count >= baseline:
            qty = current_count - baseline
        else:
            # 카운터 리셋 발생
            qty = current_count

        out[ip] = {
            "qty": max(0, qty),
            "name": str(name or ip),
        }

    return out


def query_latest_machine_snapshot():
    """
    최근 10분 데이터 중 ip별 최신 상태를 조회한다.
    대시보드 메인 목록에서 사용하는 핵심 조회 함수.
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT *
            FROM (
                SELECT *,
                       ROW_NUMBER() OVER (
                           PARTITION BY ip
                           ORDER BY [timestamp] DESC
                       ) AS rn
                FROM dbo.machine_status
                WHERE [timestamp] >= DATEADD(MINUTE, -10, GETDATE())
            ) t
            WHERE rn = 1
            ORDER BY name
        """)

        rows = rows_from_cursor_fetchall(cur, cur.fetchall())
        today_map = query_today_output_by_ip(conn, dt.datetime.now())

    for row in rows:
        ip = str(row.get("ip") or "")
        qty = today_map.get(ip, {}).get("qty", 0)

        row["part_count_today"] = qty
        row["part_count"] = qty
        row["timestamp"] = format_ts(row.get("timestamp"))

    return {
        "count": len(rows),
        "rows": rows,
        "refreshed_at": time.time(),
    }


# =========================================================
# 7. Query Functions - Replay
# =========================================================
def query_replay_range_by_date(date_str: str):
    """
    특정 날짜의 replay 가능한 최소/최대 시각을 조회한다.
    """
    target_date = dt.datetime.strptime(date_str, "%Y-%m-%d").date()
    start_dt = dt.datetime.combine(target_date, dt.time.min)
    end_dt = dt.datetime.combine(target_date, dt.time.max)

    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT MIN([timestamp]), MAX([timestamp])
            FROM dbo.machine_status
            WHERE [timestamp] BETWEEN ? AND ?
        """, (start_dt, end_dt))
        row = cur.fetchone()

    if not row or not row[0] or not row[1]:
        return None

    min_ts, max_ts = row
    return {
        "ok": True,
        "date": date_str,
        "min": min_ts.strftime("%Y-%m-%d %H:%M:%S"),
        "max": max_ts.strftime("%Y-%m-%d %H:%M:%S"),
        "step_sec": REPLAY_STEP_SEC,
    }


def query_replay_snapshot(at_time: dt.datetime):
    """
    특정 시각 기준으로 각 설비의 최신 상태를 조회한다.

    기준 시각 이전 데이터가 없는 경우,
    해당 날짜의 가장 이른 시각으로 fallback 한다.

    생산량은 target_at 기준 당일 누적 생산량으로 통일한다.
    """
    day_start = dt.datetime.combine(at_time.date(), dt.time.min)
    day_end = dt.datetime.combine(at_time.date(), dt.time.max)

    with get_conn() as conn:
        cur = conn.cursor()

        cur.execute("""
            SELECT *
            FROM (
                SELECT *,
                       ROW_NUMBER() OVER (
                           PARTITION BY ip
                           ORDER BY [timestamp] DESC
                       ) AS rn
                FROM dbo.machine_status
                WHERE [timestamp] BETWEEN ? AND ?
                  AND [timestamp] <= ?
            ) t
            WHERE rn = 1
            ORDER BY name
        """, (day_start, day_end, at_time))
        rows = rows_from_cursor_fetchall(cur, cur.fetchall())

        actual_target = at_time

        if not rows:
            cur.execute("""
                SELECT MIN([timestamp])
                FROM dbo.machine_status
                WHERE [timestamp] BETWEEN ? AND ?
            """, (day_start, day_end))
            first_row = cur.fetchone()
            first_ts = first_row[0] if first_row else None

            if first_ts:
                actual_target = first_ts

                cur.execute("""
                    SELECT *
                    FROM (
                        SELECT *,
                               ROW_NUMBER() OVER (
                                   PARTITION BY ip
                                   ORDER BY [timestamp] DESC
                               ) AS rn
                        FROM dbo.machine_status
                        WHERE [timestamp] BETWEEN ? AND ?
                          AND [timestamp] <= ?
                    ) t
                    WHERE rn = 1
                    ORDER BY name
                """, (day_start, day_end, actual_target))
                rows = rows_from_cursor_fetchall(cur, cur.fetchall())

        today_map = query_today_output_by_ip(conn, actual_target)

    for row in rows:
        ip = str(row.get("ip") or "")
        qty = today_map.get(ip, {}).get("qty", 0)

        row["part_count_today"] = qty
        row["part_count"] = qty
        row["timestamp"] = format_ts(row.get("timestamp"))

    return {
        "ok": True,
        "target_at": actual_target.strftime("%Y-%m-%d %H:%M:%S"),
        "count": len(rows),
        "rows": rows,
        "step_sec": REPLAY_STEP_SEC,
    }


# =========================================================
# 8. Query Functions - Operation Status
# =========================================================
def query_operation_status(start_date: str, end_date: str, keyword: str = ""):
    """
    조회 기간 동안 설비별 가동/비가동 시간 합계를 계산한다.

    방식:
    - 일자별, 설비별 첫 row / 마지막 row를 구한다.
    - total_operating_min / total_downtime_min 차이를 일별 delta로 본다.
    - 그 값을 기간 전체로 누적한다.
    """
    start_day = dt.datetime.strptime(start_date, "%Y-%m-%d").date()
    end_day = dt.datetime.strptime(end_date, "%Y-%m-%d").date()

    start_dt = dt.datetime.combine(start_day, dt.time.min)
    end_dt = dt.datetime.combine(end_day + dt.timedelta(days=1), dt.time.min)

    keyword = (keyword or "").strip()

    with get_conn() as conn:
        cur = conn.cursor()

        sql = """
            WITH base AS (
                SELECT
                    CAST([timestamp] AS DATE) AS work_date,
                    [timestamp],
                    [ip],
                    [name],
                    ISNULL([total_operating_min], 0) AS total_operating_min,
                    ISNULL([total_downtime_min], 0) AS total_downtime_min
                FROM dbo.machine_status
                WHERE [timestamp] >= ?
                  AND [timestamp] < ?
        """
        params = [start_dt, end_dt]

        if keyword:
            like_kw = f"{keyword}%"
            sql += """
                  AND (
                        [name] LIKE ?
                     OR [ip] LIKE ?
                  )
            """
            params.extend([like_kw, like_kw])

        sql += """
            ),
            ranked AS (
                SELECT
                    work_date,
                    [timestamp],
                    ip,
                    name,
                    total_operating_min,
                    total_downtime_min,
                    ROW_NUMBER() OVER (
                        PARTITION BY work_date, ip
                        ORDER BY [timestamp] ASC
                    ) AS rn_first,
                    ROW_NUMBER() OVER (
                        PARTITION BY work_date, ip
                        ORDER BY [timestamp] DESC
                    ) AS rn_last
                FROM base
            ),
            daily_delta AS (
                SELECT
                    f.work_date,
                    f.ip,
                    COALESCE(l.name, f.name, f.ip) AS name,
                    CASE
                        WHEN l.total_operating_min - f.total_operating_min < 0 THEN 0
                        ELSE l.total_operating_min - f.total_operating_min
                    END AS operating_min,
                    CASE
                        WHEN l.total_downtime_min - f.total_downtime_min < 0 THEN 0
                        ELSE l.total_downtime_min - f.total_downtime_min
                    END AS downtime_min
                FROM ranked f
                JOIN ranked l
                  ON f.work_date = l.work_date
                 AND f.ip = l.ip
                WHERE f.rn_first = 1
                  AND l.rn_last = 1
            )
            SELECT
                ip,
                name,
                SUM(operating_min) AS operating_min,
                SUM(downtime_min) AS downtime_min
            FROM daily_delta
            GROUP BY ip, name
            ORDER BY name, ip
        """

        cur.execute(sql, params)
        rows = cur.fetchall()

    if not rows:
        return []

    result = []
    for row in rows:
        name = str(getattr(row, "name", "") or "-")
        ip = str(getattr(row, "ip", "") or "-")
        op_min = max(0, safe_int(getattr(row, "operating_min", 0), 0))
        down_min = max(0, safe_int(getattr(row, "downtime_min", 0), 0))

        base = op_min + down_min
        rate = round((op_min / base) * 100, 1) if base > 0 else 0.0

        result.append({
            "name": name,
            "ip": ip,
            "utilization_rate": rate,
            "operating_min": op_min,
            "downtime_min": down_min,
        })

    return result


def append_operation_summary_row(rows: list):
    """
    설비 가동 현황 리스트 마지막에 합계 row를 추가한다.
    """
    total_operating = sum(safe_int(r.get("operating_min"), 0) for r in rows)
    total_downtime = sum(safe_int(r.get("downtime_min"), 0) for r in rows)

    avg_rate = 0.0
    if rows:
        avg_rate = round(
            sum(safe_float(r.get("utilization_rate"), 0) for r in rows) / len(rows),
            1
        )

    return rows + [{
        "name": "합계",
        "ip": "-",
        "utilization_rate": avg_rate,
        "operating_min": total_operating,
        "downtime_min": total_downtime,
        "is_summary": True,
    }]


def query_production_status(start_date: str, end_date: str, view_type: str = "machine", keyword: str = ""):
    """
    생산 실적 현황 조회.

    view_type:
    - machine : 설비 기준 집계
    - product : 제품 기준 집계

    기준:
    - part_name   : 품명
    - part_count  : 당일 누적 카운터로 보고, 일 생산량은 MAX - MIN
    - total_count : 설비 기준 누계
    """
    start_day = dt.datetime.strptime(start_date, "%Y-%m-%d").date()
    end_day = dt.datetime.strptime(end_date, "%Y-%m-%d").date()

    if start_day > end_day:
        raise ValueError("start_date must be <= end_date")

    start_dt = dt.datetime.combine(start_day, dt.time.min)
    end_dt = dt.datetime.combine(end_day + dt.timedelta(days=1), dt.time.min)

    view_type = (view_type or "machine").strip().lower()
    if view_type not in ("machine", "product"):
        view_type = "machine"

    keyword = (keyword or "").strip()

    with get_conn() as conn:
        cur = conn.cursor()

        base_sql = """
            WITH daily_product AS (
                SELECT
                    CAST([timestamp] AS DATE) AS work_date,
                    [ip],
                    COALESCE(NULLIF([name], ''), [ip]) AS name,
                    COALESCE(NULLIF([part_name], ''), '(미지정)') AS part_name,

                    /* 하루 생산량 = MAX(part_count) - MIN(part_count) */
                    CASE
                        WHEN MAX(ISNULL([part_count], 0)) - MIN(ISNULL([part_count], 0)) < 0 THEN 0
                        ELSE MAX(ISNULL([part_count], 0)) - MIN(ISNULL([part_count], 0))
                    END AS qty,

                    MAX(ISNULL([total_count], 0)) AS total_count
                FROM dbo.machine_status
                WHERE [timestamp] >= ?
                  AND [timestamp] < ?
        """
        params = [start_dt, end_dt]

        if keyword:
            like_kw = f"%{keyword}%"
            base_sql += """
                  AND (
                        [name] LIKE ?
                     OR [ip] LIKE ?
                     OR [part_name] LIKE ?
                  )
            """
            params.extend([like_kw, like_kw, like_kw])

        base_sql += """
                GROUP BY
                    CAST([timestamp] AS DATE),
                    [ip],
                    COALESCE(NULLIF([name], ''), [ip]),
                    COALESCE(NULLIF([part_name], ''), '(미지정)')
            )
        """

        if view_type == "machine":
            sql = base_sql + """
                , machine_totals AS (
                    SELECT
                        ip,
                        name,
                        SUM(qty) AS total_qty,
                        MAX(total_count) AS cumulative_qty,
                        COUNT(DISTINCT work_date) AS work_days
                    FROM daily_product
                    GROUP BY ip, name
                ),
                product_totals AS (
                    SELECT
                        ip,
                        name,
                        part_name,
                        SUM(qty) AS product_qty
                    FROM daily_product
                    GROUP BY ip, name, part_name
                ),
                top_product AS (
                    SELECT
                        ip,
                        name,
                        part_name,
                        product_qty,
                        ROW_NUMBER() OVER (
                            PARTITION BY ip
                            ORDER BY product_qty DESC, part_name ASC
                        ) AS rn
                    FROM product_totals
                ),
                product_breakdown AS (
                    SELECT
                        ip,
                        name,
                        STRING_AGG(
                            CONCAT(part_name, '(', CAST(product_qty AS NVARCHAR(30)), ')'),
                            ' / '
                        ) AS product_breakdown
                    FROM product_totals
                    GROUP BY ip, name
                )
                SELECT
                    mt.name,
                    COALESCE(tp.part_name, '-') AS product_name,
                    mt.total_qty,
                    mt.cumulative_qty,
                    ROUND(
                        CAST(mt.total_qty AS FLOAT) / NULLIF(mt.work_days, 0),
                        1
                    ) AS avg_qty,
                    COALESCE(pb.product_breakdown, '') AS product_breakdown
                FROM machine_totals mt
                LEFT JOIN top_product tp
                  ON mt.ip = tp.ip
                 AND mt.name = tp.name
                 AND tp.rn = 1
                LEFT JOIN product_breakdown pb
                  ON mt.ip = pb.ip
                 AND mt.name = pb.name
                ORDER BY mt.name, mt.ip
            """
        else:
            sql = base_sql + """
                , product_totals AS (
                    SELECT
                        part_name,
                        SUM(qty) AS total_qty,
                        COUNT(DISTINCT work_date) AS work_days
                    FROM daily_product
                    GROUP BY part_name
                ),
                machine_totals AS (
                    SELECT
                        part_name,
                        ip,
                        name,
                        SUM(qty) AS machine_qty,
                        ROW_NUMBER() OVER (
                            PARTITION BY part_name
                            ORDER BY SUM(qty) DESC, name ASC
                        ) AS rn
                    FROM daily_product
                    GROUP BY part_name, ip, name
                )
                SELECT
                    COALESCE(mt.name, '-') AS name,
                    pt.part_name AS product_name,
                    pt.total_qty,
                    pt.total_qty AS cumulative_qty,
                    ROUND(
                        CAST(pt.total_qty AS FLOAT) / NULLIF(pt.work_days, 0),
                        1
                    ) AS avg_qty,
                    CAST('' AS NVARCHAR(4000)) AS product_breakdown
                FROM product_totals pt
                LEFT JOIN machine_totals mt
                  ON pt.part_name = mt.part_name
                 AND mt.rn = 1
                ORDER BY pt.part_name
            """

        cur.execute(sql, params)
        rows = cur.fetchall()

    result = []
    for row in rows:
        raw_breakdown = str(getattr(row, "product_breakdown", "") or "").strip()

        breakdown_rows = []
        if raw_breakdown:
            for item in raw_breakdown.split(" / "):
                item = item.strip()
                if not item:
                    continue

                # 예: "(HPV102 HEAD BLOCK)(61)" 또는 "UNKNOWN(3006)"
                if "(" in item and item.endswith(")"):
                    pos = item.rfind("(")
                    pname = item[:pos].strip()
                    qty_text = item[pos + 1:-1].strip()

                    breakdown_rows.append({
                        "product_name": pname or "-",
                        "qty": safe_int(qty_text, 0),
                    })

        result.append({
            "name": str(getattr(row, "name", "") or "-"),
            "product_name": str(getattr(row, "product_name", "") or "-"),
            "total_qty": max(0, safe_int(getattr(row, "total_qty", 0), 0)),
            "cumulative_qty": max(0, safe_int(getattr(row, "cumulative_qty", 0), 0)),
            "avg_qty": round(safe_float(getattr(row, "avg_qty", 0), 0), 1),
            "product_breakdown": raw_breakdown,
            "breakdown_rows": breakdown_rows,
        })

    return result


def append_production_summary_row(rows: list, view_type: str = "machine"):
    """
    생산 실적 현황 리스트 마지막에 합계 row를 추가한다.
    """
    total_qty = sum(safe_int(r.get("total_qty"), 0) for r in rows)
    cumulative_qty = sum(safe_int(r.get("cumulative_qty"), 0) for r in rows)

    avg_qty = 0.0
    if rows:
        avg_qty = round(
            sum(safe_float(r.get("avg_qty"), 0) for r in rows) / len(rows),
            1
        )

    return rows + [{
        "name": "합계",
        "product_name": "-",
        "total_qty": total_qty,
        "cumulative_qty": cumulative_qty,
        "avg_qty": avg_qty,
        "product_breakdown": "",
        "is_summary": True,
        "view_type": view_type,
    }]


def build_production_status_excel(start_date: str, end_date: str, view_type: str = "machine", keyword: str = ""):
    """
    생산 실적 현황 데이터를 xlsx 파일로 생성한다.
    """
    rows = query_production_status(
        start_date=start_date,
        end_date=end_date,
        view_type=view_type,
        keyword=keyword,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "생산실적현황"

    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="EAF2F8")

    ws["A1"] = "생산 실적 현황"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = "조회기간"
    ws["B2"] = f"{start_date} ~ {end_date}"
    ws["A3"] = "조회유형"
    ws["B3"] = "설비" if view_type == "machine" else "제품"
    ws["A4"] = "키워드"
    ws["B4"] = keyword or "-"

    if view_type == "machine":
        headers = ["설비명", "품명", "총계", "누계", "평균", "품목별 누계"]
    else:
        headers = ["설비명", "품명", "총계", "누계", "평균"]

    start_row = 6

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = bold
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, row in enumerate(rows, start=start_row + 1):
        ws.cell(i, 1, row["name"])
        ws.cell(i, 2, row["product_name"])
        ws.cell(i, 3, row["total_qty"])
        ws.cell(i, 4, row["cumulative_qty"])
        ws.cell(i, 5, row["avg_qty"])

        if view_type == "machine":
            ws.cell(i, 6, row["product_breakdown"])

    widths = {
        "A": 18,
        "B": 30,
        "C": 12,
        "D": 12,
        "E": 12,
        "F": 50,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def query_tool_status(start_date: str, end_date: str, machine: str = "", keyword: str = ""):
    """
    공구관리 현황 조회.

    개선 사항:
    - 최대 조회 기간 30일 제한
    - 10초 캐시
    - ROW_NUMBER 제거
    - GROUP BY + MAX(timestamp) 방식 사용
    """
    start_d = parse_date_only(start_date)
    end_d = parse_date_only(end_date)

    if not start_d or not end_d:
        raise ValueError("invalid start_date / end_date")

    if start_d > end_d:
        raise ValueError("start_date must be <= end_date")

    # 최대 조회 기간 제한 (30일)
    day_span = (end_d - start_d).days + 1
    if day_span > 30:
        raise ValueError("공구관리 현황 조회 기간은 최대 30일까지 가능합니다.")

    machine = (machine or "").strip()
    keyword = (keyword or "").strip()

    # 캐시 확인 (10초)
    cache_key = make_tool_cache_key(start_date, end_date, machine, keyword)
    cached = _TOOL_STATUS_CACHE.get(cache_key)
    now_ts = time.time()
    if cached and (now_ts - cached["ts"]) <= 10:
        return cached["rows"]

    start_dt = dt.datetime.combine(start_d, dt.time.min)
    end_dt = dt.datetime.combine(end_d + dt.timedelta(days=1), dt.time.min)

    with get_conn() as conn:
        cols = get_machine_status_columns(conn)

        ts_col = pick_existing_column(cols, ["timestamp"])
        ip_col = pick_existing_column(cols, ["ip", "machine_ip"])
        name_col = pick_existing_column(cols, ["name", "machine_name"])
        status_col = pick_existing_column(cols, ["status", "op_status", "operation_status"])
        tool_col = pick_existing_column(cols, ["tool_no", "tool_number", "tool"])
        use_time_col = pick_existing_column(cols, [
            "tool_used",
            "use_time",
            "used_time",
            "tool_use_time",
            "tool_used_time",
            "current_use_time",
            "tool_current_use_time",
        ])
        limit_time_col = pick_existing_column(cols, [
            "tool_limit",
            "limit_time",
            "tool_limit_time",
            "tool_life_limit",
            "max_tool_life",
            "tool_max_life",
            "life_limit",
        ])
        life_pct_col = pick_existing_column(cols, [
            "tool_life_pct",
            "tool_life",
            "life_pct",
            "life_rate",
        ])
        warning_col = pick_existing_column(cols, [
            "warning",
            "tool_warning",
            "alarm_msg",
            "tool_alarm_msg",
        ])
        product_col = pick_existing_column(cols, [
            "part_name",
            "product_name",
            "item_name",
            "pname",
            "품명",
        ])

        if not ts_col or not ip_col or not tool_col:
            raise RuntimeError(
                "machine_status 테이블에 timestamp / ip / tool_no 계열 컬럼이 필요합니다."
            )

        ts_sql = sql_ident(ts_col)
        ip_sql = sql_ident(ip_col)
        name_sql = sql_ident(name_col) if name_col else None
        status_sql = sql_ident(status_col) if status_col else None
        tool_sql = sql_ident(tool_col)
        use_time_sql = sql_ident(use_time_col) if use_time_col else None
        limit_time_sql = sql_ident(limit_time_col) if limit_time_col else None
        life_pct_sql = sql_ident(life_pct_col) if life_pct_col else None
        warning_sql = sql_ident(warning_col) if warning_col else None
        product_sql = sql_ident(product_col) if product_col else None

        name_expr = (
            f"COALESCE(CAST(ms.{name_sql} AS NVARCHAR(200)), CAST(ms.{ip_sql} AS NVARCHAR(100)))"
            if name_sql else
            f"CAST(ms.{ip_sql} AS NVARCHAR(100))"
        )

        status_expr = (
            f"CAST(ms.{status_sql} AS NVARCHAR(50))"
            if status_sql else
            "CAST('UNKNOWN' AS NVARCHAR(50))"
        )

        use_time_expr = (
            f"COALESCE(TRY_CONVERT(FLOAT, ms.{use_time_sql}), 0)"
            if use_time_sql else
            "0"
        )

        limit_time_expr = (
            f"COALESCE(TRY_CONVERT(FLOAT, ms.{limit_time_sql}), 0)"
            if limit_time_sql else
            "0"
        )

        life_pct_expr = (
            f"COALESCE(TRY_CONVERT(FLOAT, ms.{life_pct_sql}), 0)"
            if life_pct_sql else
            "0"
        )

        warning_expr = (
            f"CAST(ms.{warning_sql} AS NVARCHAR(200))"
            if warning_sql else
            "CAST('' AS NVARCHAR(200))"
        )

        product_expr = (
            f"CAST(ms.{product_sql} AS NVARCHAR(300))"
            if product_sql else
            "CAST('' AS NVARCHAR(300))"
        )

        current_tool_expr = f"TRY_CONVERT(INT, latest_machine.{tool_sql})"

        sql = f"""
            WITH latest_tool_ts AS (
                SELECT
                    CAST({ip_sql} AS NVARCHAR(100)) AS ip,
                    TRY_CONVERT(INT, {tool_sql}) AS tool_no,
                    MAX({ts_sql}) AS last_ts
                FROM dbo.machine_status
                WHERE {ts_sql} >= ?
                  AND {ts_sql} < ?
                  AND {ip_sql} IS NOT NULL
                  AND {tool_sql} IS NOT NULL
                  AND TRY_CONVERT(INT, {tool_sql}) IS NOT NULL
                  AND TRY_CONVERT(INT, {tool_sql}) <> 0
                GROUP BY
                    CAST({ip_sql} AS NVARCHAR(100)),
                    TRY_CONVERT(INT, {tool_sql})
            ),
            latest_machine_ts AS (
                SELECT
                    CAST({ip_sql} AS NVARCHAR(100)) AS ip,
                    MAX({ts_sql}) AS last_ts
                FROM dbo.machine_status
                WHERE {ts_sql} < ?
                  AND {ip_sql} IS NOT NULL
                GROUP BY CAST({ip_sql} AS NVARCHAR(100))
            )
            SELECT
                {name_expr} AS name,
                CAST(ms.{ip_sql} AS NVARCHAR(100)) AS ip,
                {status_expr} AS status,
                TRY_CONVERT(INT, ms.{tool_sql}) AS tool_no,
                {use_time_expr} AS used_time,
                {limit_time_expr} AS limit_time,
                {life_pct_expr} AS life_pct,
                {warning_expr} AS warning_text,
                {product_expr} AS product_name,
                ms.{ts_sql} AS last_used,
                CASE
                    WHEN {current_tool_expr} = TRY_CONVERT(INT, ms.{tool_sql}) THEN 1
                    ELSE 0
                END AS is_current_use
            FROM latest_tool_ts lt
            JOIN dbo.machine_status ms
              ON CAST(ms.{ip_sql} AS NVARCHAR(100)) = lt.ip
             AND TRY_CONVERT(INT, ms.{tool_sql}) = lt.tool_no
             AND ms.{ts_sql} = lt.last_ts
            LEFT JOIN latest_machine_ts lmt
              ON lmt.ip = lt.ip
            LEFT JOIN dbo.machine_status latest_machine
              ON CAST(latest_machine.{ip_sql} AS NVARCHAR(100)) = lmt.ip
             AND latest_machine.{ts_sql} = lmt.last_ts
            WHERE 1 = 1
        """

        params = [start_dt, end_dt, end_dt]

        if machine and machine != "전체":
            if name_sql:
                sql += f"""
                  AND (
                        CAST(ms.{ip_sql} AS NVARCHAR(100)) = ?
                     OR CAST(ms.{name_sql} AS NVARCHAR(200)) = ?
                  )
                """
            else:
                sql += f"""
                  AND CAST(ms.{ip_sql} AS NVARCHAR(100)) = ?
                """

            if name_sql:
                params.extend([machine, machine])
            else:
                params.append(machine)

        if keyword:
            sql += """
              AND product_name LIKE ?
            """
            params.append(f"%{keyword}%")

        sql += """
            ORDER BY name, ip, tool_no
        """

        cur = conn.cursor()
        cur.execute(sql, params)
        db_rows = cur.fetchall()

    result = []
    for row in db_rows:
        name = str(getattr(row, "name", "") or "-")
        ip = str(getattr(row, "ip", "") or "-")
        status = str(getattr(row, "status", "") or "UNKNOWN").upper()
        tool_no = safe_int(getattr(row, "tool_no", 0), 0)

        used_time = safe_float(getattr(row, "used_time", 0), 0)
        limit_time = safe_float(getattr(row, "limit_time", 0), 0)
        life_pct = safe_float(getattr(row, "life_pct", 0), 0)

        if life_pct <= 0 and limit_time > 0 and used_time > 0:
            life_pct = round((used_time / limit_time) * 100, 1)

        warning_text = "경고" if safe_int(getattr(row, "tool_warn", 0), 0) == 1 else ""

        last_used = format_ts(getattr(row, "last_used", None))
        product_name = str(getattr(row, "product_name", "") or "")
        is_current_use = safe_int(getattr(row, "is_current_use", 0), 0) == 1

        result.append({
            "name": name,
            "ip": ip,
            "status": status,
            "tool_no": tool_no,
            "used_time": int(round(used_time)),
            "limit_time": int(round(limit_time)),
            "life_pct": int(round(life_pct)),
            "warning": warning_text,
            "product_name": product_name,
            "last_used": last_used,
            "is_current_use": is_current_use,
            "current_use": "사용중" if is_current_use else "",
        })

    # 캐시 저장
    _TOOL_STATUS_CACHE[cache_key] = {
        "ts": time.time(),
        "rows": result,
    }

    # 캐시가 너무 커지는 것 방지
    if len(_TOOL_STATUS_CACHE) > 100:
        old_keys = sorted(
            _TOOL_STATUS_CACHE.keys(),
            key=lambda k: _TOOL_STATUS_CACHE[k]["ts"]
        )
        for k in old_keys[:20]:
            _TOOL_STATUS_CACHE.pop(k, None)

    return result


# =========================================================
# 9. Query Functions - Machine / Tool Detail
# =========================================================
def query_machine_row_at(ip: str, at_time: dt.datetime):
    """
    특정 시각 이전 기준으로 해당 설비의 최신 row 1건 조회.
    replay 기준 당일 누적 생산량을 함께 반영한다.
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT TOP 1 *
            FROM dbo.machine_status
            WHERE ip = ?
              AND [timestamp] <= ?
            ORDER BY [timestamp] DESC, id DESC
        """, (ip, at_time))
        row = row_from_cursor_fetchone(cur, cur.fetchone())

        if row:
            today_map = query_today_output_by_ip(conn, at_time)
            qty = today_map.get(str(ip), {}).get("qty", 0)
            row["part_count_today"] = qty
            row["part_count"] = qty
            row["timestamp"] = format_ts(row.get("timestamp"))

    return row


def query_latest_tool_row(ip: str, tool_no: int):
    """
    live 모드에서 특정 설비/공구의 최신 row 조회.
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT TOP 1 *
            FROM dbo.machine_status
            WHERE ip = ?
              AND tool_no = ?
            ORDER BY [timestamp] DESC
        """, (ip, tool_no))
        row = row_from_cursor_fetchone(cur, cur.fetchone())

    if row:
        row["timestamp"] = format_ts(row.get("timestamp"))
    return row


def query_tool_row_at(ip: str, tool_no: int, at_time: dt.datetime):
    """
    replay 모드에서 특정 설비/공구의 기준 시각 이전 최신 row 조회.
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT TOP 1 *
            FROM dbo.machine_status
            WHERE ip = ?
              AND tool_no = ?
              AND [timestamp] <= ?
            ORDER BY [timestamp] DESC
        """, (ip, tool_no, at_time))
        row = row_from_cursor_fetchone(cur, cur.fetchone())

    if row:
        row["timestamp"] = format_ts(row.get("timestamp"))
    return row


def query_machine_tools(ip: str):
    """
    특정 설비에서 사용된 tool_no 목록 조회.
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT tool_no
            FROM dbo.machine_status
            WHERE ip = ?
              AND tool_no IS NOT NULL
              AND tool_no <> 0
            ORDER BY tool_no
        """, (ip,))
        return [int(r[0]) for r in cur.fetchall()]


def query_tool_chart_rows(ip: str, tool_no: int, limit: int, at_time=None):
    """
    공구 차트용 시계열 조회.

    live:
        ip + tool_no 최신 N건
    replay:
        ip + tool_no + at_time 이하 최신 N건
    """
    with get_conn() as conn:
        cur = conn.cursor()

        if at_time is None:
            cur.execute(f"""
                SELECT TOP {limit} [timestamp], utilization_rate, rpm
                FROM dbo.machine_status
                WHERE ip = ?
                  AND tool_no = ?
                ORDER BY [timestamp] DESC
            """, (ip, tool_no))
        else:
            cur.execute(f"""
                SELECT TOP {limit} [timestamp], utilization_rate, rpm
                FROM dbo.machine_status
                WHERE ip = ?
                  AND tool_no = ?
                  AND [timestamp] <= ?
                ORDER BY [timestamp] DESC
            """, (ip, tool_no, at_time))

        rows = cur.fetchall()

    return list(reversed(rows))


# =========================================================
# 10. Payload Builders
# =========================================================
def build_machine_current_payload(ip: str, row: dict):
    """
    설비 현재 상태 응답 포맷을 통일한다.
    """
    return {
        "ok": True,
        "ip": ip,
        "name": row.get("name") or ip,
        "status": row.get("status") or "UNKNOWN",
        "timestamp": row.get("timestamp"),
        "currentToolNo": safe_int(row.get("tool_no"), 0),
        "loadPct": safe_float(row.get("utilization_rate"), 0),
        "rpm": safe_float(row.get("rpm"), 0),
        "prodMin": safe_int(row.get("total_operating_min") or row.get("operating_min"), 0),
        "program": row.get("onum"),
        "alarm": row.get("alarm", 0),
        "partCount": safe_int(row.get("part_count_today") or row.get("part_count"), 0),
    }


def build_tool_detail_payload(ip: str, current: dict, selected_tool_row: dict, tool_no: int, mode: str, replay_at=None):
    """
    공구 상세 응답 payload를 live/replay 공통으로 조립한다.

    규칙:
    - 요청한 tool_no가 현재 장착 공구면 현재 설비의 load/rpm 사용
    - 아니면 선택 공구 row의 load만 사용, rpm은 0 처리
    """
    current_tool_no = safe_int(current.get("tool_no"), 0)
    is_current_tool = (tool_no == current_tool_no)

    payload = {
        "ok": True,
        "ip": ip,
        "name": current.get("name") or ip,
        "status": current.get("status") or "UNKNOWN",
        "timestamp": current.get("timestamp"),
        "mode": mode,
        "toolNo": tool_no,
        "currentToolNo": current_tool_no,
        "isCurrentTool": is_current_tool,
        "loadPct": safe_float(
            current.get("utilization_rate") if is_current_tool else selected_tool_row.get("utilization_rate"),
            0
        ),
        "rpm": safe_float(current.get("rpm"), 0) if is_current_tool else 0,
        "prodMin": safe_int(current.get("total_operating_min") or current.get("operating_min"), 0),
        "program": current.get("onum"),
        "alarm": current.get("alarm", 0),
        "machineToolNo": current_tool_no,
    }

    if replay_at:
        payload["replayAt"] = replay_at.strftime("%Y-%m-%d %H:%M:%S")

    return payload


def build_tool_chart_payload(ip: str, tool_no: int, current_tool_no: int, rows, mode: str, replay_at=None):
    """
    공구 차트 응답 payload 조립.
    """
    is_current_tool = (tool_no == current_tool_no)

    labels = []
    load = []
    rpm = []

    for ts, util, row_rpm in rows:
        labels.append(ts.strftime("%H:%M:%S") if hasattr(ts, "strftime") else str(ts))
        load.append(round(safe_float(util, 0), 2))
        rpm.append(round(safe_float(row_rpm, 0), 2) if is_current_tool else 0)

    payload = {
        "ok": True,
        "ip": ip,
        "toolNo": tool_no,
        "currentToolNo": current_tool_no,
        "isCurrentTool": is_current_tool,
        "mode": mode,
        "labels": labels,
        "load": load,
        "rpm": rpm,
    }

    if replay_at:
        payload["replayAt"] = replay_at.strftime("%Y-%m-%d %H:%M:%S")

    return payload


# =========================================================
# 11. Excel Builder
# =========================================================
def build_operation_status_excel(start_date: str, end_date: str, keyword: str = ""):
    """
    설비가동현황 데이터를 xlsx 파일로 생성한다.
    """
    rows = query_operation_status(start_date, end_date, keyword)

    wb = Workbook()
    ws = wb.active
    ws.title = "설비가동현황"

    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="EAF2F8")

    ws["A1"] = "설비가동 현황"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = "조회기간"
    ws["B2"] = f"{start_date} ~ {end_date}"
    ws["A3"] = "키워드"
    ws["B3"] = keyword or "-"

    headers = ["설비명", "IP", "가동율(%)", "가동(분)", "비가동(분)"]
    start_row = 5

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = bold
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, row in enumerate(rows, start=start_row + 1):
        ws.cell(i, 1, row["name"])
        ws.cell(i, 2, row["ip"])
        ws.cell(i, 3, row["utilization_rate"])
        ws.cell(i, 4, row["operating_min"])
        ws.cell(i, 5, row["downtime_min"])

    for col in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 18

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def build_tool_status_excel(start_date: str, end_date: str, machine: str = "", keyword: str = ""):
    """
    공구관리 현황 데이터를 xlsx 파일로 생성한다.
    """
    rows = query_tool_status(
        start_date=start_date,
        end_date=end_date,
        machine=machine,
        keyword=keyword,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "공구관리현황"

    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="EAF2F8")

    ws["A1"] = "공구관리 현황"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = "조회기간"
    ws["B2"] = f"{start_date} ~ {end_date}"
    ws["A3"] = "설비"
    ws["B3"] = machine or "전체"
    ws["A4"] = "품명 키워드"
    ws["B4"] = keyword or "-"

    headers = [
        "설비명",
        "IP",
        "상태",
        "공구번호",
        "사용시간",
        "한계시간",
        "수명(%)",
        "경고",
        "제품명",
        "마지막사용",
        "현재사용",
    ]

    start_row = 6

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = bold
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, row in enumerate(rows, start=start_row + 1):
        ws.cell(i, 1, row["name"])
        ws.cell(i, 2, row["ip"])
        ws.cell(i, 3, row["status"])
        ws.cell(i, 4, row["tool_no"])
        ws.cell(i, 5, row["used_time"])
        ws.cell(i, 6, row["limit_time"])
        ws.cell(i, 7, row["life_pct"])
        ws.cell(i, 8, row["warning"])
        ws.cell(i, 9, row["product_name"])
        ws.cell(i, 10, row["last_used"])
        ws.cell(i, 11, row["current_use"])

    widths = {
        "A": 16,
        "B": 18,
        "C": 10,
        "D": 10,
        "E": 12,
        "F": 12,
        "G": 10,
        "H": 12,
        "I": 32,
        "J": 20,
        "K": 12,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# =========================================================
# 12. Background Worker
# =========================================================
def refresh_loop():
    """
    최신 설비 스냅샷을 주기적으로 갱신한다.
    """
    while True:
        t0 = time.time()
        try:
            _CACHE["data"] = query_latest_machine_snapshot()
            _CACHE["ts"] = time.time()
            _CACHE["error"] = None
        except Exception as e:
            _CACHE["error"] = f"{type(e).__name__}: {e}"

        time.sleep(max(0.1, REFRESH_SEC - (time.time() - t0)))


def start_worker_once():
    """
    워커 스레드를 1회만 시작한다.
    debug reloader 중복 실행을 방지한다.
    """
    if _CACHE["running"]:
        return

    if app.debug and os.environ.get("WERKZEUG_RUN_MAIN") != "true":
        return

    _CACHE["running"] = True
    threading.Thread(target=refresh_loop, daemon=True).start()


@app.before_request
def ensure_worker():
    """
    첫 요청 시점에 워커 실행 보장.
    """
    start_worker_once()


# =========================================================
# 13. Static / Page Routes
# =========================================================
@app.get("/assets/<path:filename>")
def assets(filename):
    """정적 assets 파일 제공."""
    return send_from_directory("assets", filename, as_attachment=False)


@app.get("/")
def home():
    """루트 접근 시 대시보드로 이동."""
    return redirect("/dashboard")


@app.get("/dashboard")
def dashboard():
    """대시보드 페이지."""
    return render_template("index.html")


@app.get("/emulator")
def emulator_page():
    """에뮬레이터 페이지."""
    return render_template("emulator.html", replay_step_sec=REPLAY_STEP_SEC)


@app.get("/machine/<path:ip>")
def machine_detail(ip):
    """설비 상세 페이지."""
    return render_template("machine_detail.html", ip=ip)


# =========================================================
# 14. Common API Routes
# =========================================================
@app.get("/health")
def health():
    """서버 및 캐시 상태 확인."""
    return jsonify({
        "ok": True,
        "cache_ready": _CACHE["data"] is not None,
        "cache_age_sec": None if not _CACHE["data"] else round(time.time() - _CACHE["ts"], 3),
        "error": _CACHE["error"],
    })


@app.get("/api/machines")
def api_machines():
    """캐시된 최신 설비 목록 반환."""
    if not _CACHE["data"]:
        return jsonify({"error": "warming up"}), 503
    return jsonify(_CACHE["data"])


# =========================================================
# 15. Replay API Routes
# =========================================================
@app.get("/api/replay/range")
def api_replay_range():
    """특정 날짜의 replay 가능 시간 범위 조회."""
    date_str = request.args.get("date", "").strip()
    if not date_str:
        return jsonify({"ok": False, "message": "date is required"}), 400

    try:
        data = query_replay_range_by_date(date_str)
    except Exception as e:
        return jsonify({"ok": False, "message": f"range query failed: {e}"}), 500

    if not data:
        return jsonify({"ok": False, "message": "해당 날짜의 기록이 없습니다."}), 404

    return jsonify(data)


@app.get("/api/replay/snapshot")
def api_replay_snapshot():
    """특정 시각 기준 replay 스냅샷 조회."""
    at_str = request.args.get("at", "").strip()
    if not at_str:
        return jsonify({"ok": False, "message": "at is required"}), 400

    at_time = parse_local_datetime(at_str)
    if not at_time:
        return jsonify({"ok": False, "message": "invalid datetime format"}), 400

    try:
        data = query_replay_snapshot(at_time)
    except Exception as e:
        return jsonify({"ok": False, "message": f"snapshot query failed: {e}"}), 500

    return jsonify(data)


# =========================================================
# 16. Operation Status API Routes
# =========================================================
@app.get("/api/operation-status")
def api_operation_status():
    """설비 가동 현황 조회."""
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    keyword = request.args.get("keyword", "").strip()

    if not start_date or not end_date:
        return jsonify({"ok": False, "message": "start_date, end_date are required"}), 400

    try:
        rows = query_operation_status(start_date, end_date, keyword)
    except Exception as e:
        return jsonify({"ok": False, "message": f"operation status query failed: {e}"}), 500

    return jsonify({
        "ok": True,
        "count": len(rows),
        "rows": rows,
        "start_date": start_date,
        "end_date": end_date,
        "keyword": keyword,
    })


@app.get("/api/operation-status/export")
def api_operation_status_export():
    """설비 가동 현황 엑셀 다운로드."""
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    keyword = request.args.get("keyword", "").strip()

    if not start_date or not end_date:
        return jsonify({"ok": False, "message": "start_date, end_date are required"}), 400

    try:
        bio = build_operation_status_excel(start_date, end_date, keyword)
    except Exception as e:
        return jsonify({"ok": False, "message": f"operation status export failed: {e}"}), 500

    filename = f"operation_status_{start_date}_{end_date}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/tool-status/machines")
def api_tool_status_machines():
    """
    공구관리 현황 설비 드롭다운 목록.
    """
    try:
        rows = query_tool_status_machine_list()
    except Exception as e:
        return jsonify({"ok": False, "message": f"machine list query failed: {e}"}), 500

    return jsonify({
        "ok": True,
        "count": len(rows),
        "rows": rows,
    })


@app.get("/api/tool-status")
def api_tool_status():
    """
    공구관리 현황 조회.

    query params:
    - start_date=YYYY-MM-DD
    - end_date=YYYY-MM-DD
    - machine=전체 또는 설비명/IP
    - keyword=품명 키워드
    """
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    machine = request.args.get("machine", "").strip()
    keyword = request.args.get("keyword", "").strip()

    if not start_date or not end_date:
        return jsonify({"ok": False, "message": "start_date, end_date are required"}), 400

    try:
        rows = query_tool_status(
            start_date=start_date,
            end_date=end_date,
            machine=machine,
            keyword=keyword,
        )
    except Exception as e:
        return jsonify({"ok": False, "message": f"tool status query failed: {e}"}), 500

    return jsonify({
        "ok": True,
        "count": len(rows),
        "rows": rows,
        "start_date": start_date,
        "end_date": end_date,
        "machine": machine or "전체",
        "keyword": keyword,
    })


@app.get("/api/tool-status/export")
def api_tool_status_export():
    """
    공구관리 현황 엑셀 다운로드.
    """
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    machine = request.args.get("machine", "").strip()
    keyword = request.args.get("keyword", "").strip()

    if not start_date or not end_date:
        return jsonify({"ok": False, "message": "start_date, end_date are required"}), 400

    try:
        bio = build_tool_status_excel(
            start_date=start_date,
            end_date=end_date,
            machine=machine,
            keyword=keyword,
        )
    except Exception as e:
        return jsonify({"ok": False, "message": f"tool status export failed: {e}"}), 500

    filename = f"tool_status_{start_date}_{end_date}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.get("/api/production-status")
def api_production_status():
    """
    생산 실적 현황 조회.

    query params:
    - start_date=YYYY-MM-DD
    - end_date=YYYY-MM-DD
    - view_type=machine|product
    - keyword=설비명/IP/품명 키워드
    """
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    view_type = request.args.get("view_type", "machine").strip().lower()
    keyword = request.args.get("keyword", "").strip()

    if not start_date or not end_date:
        return jsonify({"ok": False, "message": "start_date, end_date are required"}), 400

    try:
        rows = query_production_status(
            start_date=start_date,
            end_date=end_date,
            view_type=view_type,
            keyword=keyword,
        )
    except Exception as e:
        return jsonify({"ok": False, "message": f"production status query failed: {e}"}), 500

    return jsonify({
        "ok": True,
        "view_type": view_type,
        "count": len(rows),
        "rows": rows,
        "start_date": start_date,
        "end_date": end_date,
        "keyword": keyword,
    })


@app.get("/api/production-status/export")
def api_production_status_export():
    """
    생산 실적 현황 엑셀 다운로드.
    """
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    view_type = request.args.get("view_type", "machine").strip().lower()
    keyword = request.args.get("keyword", "").strip()

    if not start_date or not end_date:
        return jsonify({"ok": False, "message": "start_date, end_date are required"}), 400

    try:
        bio = build_production_status_excel(
            start_date=start_date,
            end_date=end_date,
            view_type=view_type,
            keyword=keyword,
        )
    except Exception as e:
        return jsonify({"ok": False, "message": f"production status export failed: {e}"}), 500

    filename = f"production_status_{view_type}_{start_date}_{end_date}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# =========================================================
# 17. Machine / Tool API Routes
# =========================================================
@app.get("/api/machine/<path:ip>/current")
def api_machine_current(ip):
    """
    설비 현재 상태 조회.
    - at 파라미터가 있으면 replay
    - 없으면 live
    """
    at_str = request.args.get("at", "").strip()

    if at_str:
        at_time = parse_local_datetime(at_str)
        if not at_time:
            return jsonify({"ok": False, "message": "invalid datetime format"}), 400

        row = query_machine_row_at(ip, at_time)
        if not row:
            return jsonify({"ok": False, "message": "machine not found at replay time"}), 404

        payload = build_machine_current_payload(ip, row)
        payload["mode"] = "replay"
        payload["replayAt"] = at_time.strftime("%Y-%m-%d %H:%M:%S")
        return jsonify(payload)

    row = get_cached_machine_row(ip)
    if not row:
        return jsonify({"ok": False, "message": "machine not found"}), 404

    payload = build_machine_current_payload(ip, row)
    payload["mode"] = "live"
    return jsonify(payload)


@app.get("/api/machine/<path:ip>/tools")
def api_machine_tools(ip):
    """해당 설비의 tool 목록 조회."""
    tools = query_machine_tools(ip)
    return jsonify({"ip": ip, "tools": tools})


@app.get("/api/machine/<path:ip>/tool/<int:tool_no>")
def api_machine_tool_detail(ip, tool_no):
    """
    특정 공구 상세 조회.
    live / replay 분기만 처리하고,
    실제 payload 조립은 builder 함수에 맡긴다.
    """
    at_str = request.args.get("at", "").strip()

    if at_str:
        at_time = parse_local_datetime(at_str)
        if not at_time:
            return jsonify({"ok": False, "message": "invalid datetime format"}), 400

        current = query_machine_row_at(ip, at_time)
        if not current:
            return jsonify({"ok": False, "message": "machine not found at replay time"}), 404

        selected_tool_row = query_tool_row_at(ip, tool_no, at_time)
        if not selected_tool_row:
            return jsonify({"ok": False, "message": "tool not found at replay time"}), 404

        payload = build_tool_detail_payload(
            ip=ip,
            current=current,
            selected_tool_row=selected_tool_row,
            tool_no=tool_no,
            mode="replay",
            replay_at=at_time,
        )
        return jsonify(payload)

    current = get_cached_machine_row(ip)
    if not current:
        return jsonify({"ok": False, "message": "machine not found"}), 404

    selected_tool_row = query_latest_tool_row(ip, tool_no)
    if not selected_tool_row:
        return jsonify({"ok": False, "message": "tool not found"}), 404

    payload = build_tool_detail_payload(
        ip=ip,
        current=current,
        selected_tool_row=selected_tool_row,
        tool_no=tool_no,
        mode="live",
    )
    return jsonify(payload)


@app.get("/api/machine/<path:ip>/tool/<int:tool_no>/chart")
def api_machine_tool_chart(ip, tool_no):
    """
    특정 공구 차트 조회.
    """
    at_str = request.args.get("at", "").strip()

    try:
        limit = int(request.args.get("limit", 60))
    except Exception:
        limit = 60

    limit = max(10, min(limit, 200))

    if at_str:
        at_time = parse_local_datetime(at_str)
        if not at_time:
            return jsonify({"ok": False, "message": "invalid datetime format"}), 400

        current = query_machine_row_at(ip, at_time)
        if not current:
            return jsonify({"ok": False, "message": "machine not found at replay time"}), 404

        current_tool_no = safe_int(current.get("tool_no"), 0)
        rows = query_tool_chart_rows(ip, tool_no, limit, at_time=at_time)

        payload = build_tool_chart_payload(
            ip=ip,
            tool_no=tool_no,
            current_tool_no=current_tool_no,
            rows=rows,
            mode="replay",
            replay_at=at_time,
        )
        return jsonify(payload)

    current = get_cached_machine_row(ip)
    if not current:
        return jsonify({"ok": False, "message": "machine not found"}), 404

    current_tool_no = safe_int(current.get("tool_no"), 0)
    rows = query_tool_chart_rows(ip, tool_no, limit)

    payload = build_tool_chart_payload(
        ip=ip,
        tool_no=tool_no,
        current_tool_no=current_tool_no,
        rows=rows,
        mode="live",
    )
    return jsonify(payload)


# =========================================================
# 18. Run
# =========================================================
if __name__ == "__main__":
    app.run(host="127.0.0.1", port=8000, debug=True)
